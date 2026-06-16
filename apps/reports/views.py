"""
Módulo de Reportes - exportación PDF, CSV y Excel
"""
import io
import csv
from datetime import datetime
from django.db.models import Q
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from apps.authentication.permissions import EsMedico
from apps.etl.models import Paciente, HistorialETL
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _filtrar_pacientes(request):
    qs = Paciente.objects.all()
    riesgo = request.query_params.get('riesgo')
    sexo = request.query_params.get('sexo')
    critico = request.query_params.get('critico')
    busqueda = request.query_params.get('busqueda')
    if riesgo:
        qs = qs.filter(riesgo_enfermedad=riesgo)
    if sexo:
        qs = qs.filter(sexo=sexo)
    if critico == 'true':
        qs = qs.filter(es_critico=True)
    if busqueda:
        qs = qs.filter(
            Q(nombres__icontains=busqueda) |
            Q(apellidos__icontains=busqueda) |
            Q(diagnostico_preliminar__icontains=busqueda)
        )
    return qs


@api_view(['GET'])
@permission_classes([IsAuthenticated, EsMedico])
def historial_etl_reporte(request):
    from rest_framework.response import Response
    from apps.etl.serializers import HistorialETLSerializer
    data = HistorialETLSerializer(HistorialETL.objects.all()[:50], many=True).data
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_pdf(request):
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        name='DocTitle',
        parent=styles['Heading1'],
        fontSize=22,
        leading=26,
        textColor=colors.HexColor('#2d0a57'),
        spaceAfter=4
    )

    subtitle_style = ParagraphStyle(
        name='DocSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#7c3aed'),
        spaceAfter=15,
    )

    meta_style = ParagraphStyle(
        name='DocMeta',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#6b6489'),
        spaceAfter=12
    )

    section_title_style = ParagraphStyle(
        name='SectionTitle',
        parent=styles['Heading2'],
        fontSize=11,
        leading=15,
        textColor=colors.HexColor('#4a1080'),
        spaceBefore=18,
        spaceAfter=8,
        fontName='Helvetica-Bold'
    )

    table_cell_style = ParagraphStyle(
        name='TableCell',
        parent=styles['Normal'],
        fontSize=7.5,
        leading=10,
        textColor=colors.HexColor('#1e1b29')
    )

    table_cell_header_style = ParagraphStyle(
        name='TableCellHeader',
        parent=styles['Normal'],
        fontSize=7.5,
        leading=10,
        textColor=colors.white,
        fontName='Helvetica-Bold'
    )

    purple = colors.HexColor('#4a1080')
    purple_light = colors.HexColor('#ede9fe')
    purple_mid = colors.HexColor('#7c3aed')
    gray_bg = colors.HexColor('#f8f7fc')
    border_color = colors.HexColor('#ddd6fe')

    RISK_COLORS = {
        'bajo':   '#0f766e',
        'medio':  '#92400e',
        'alto':   '#9f1239',
        'critico': '#be123c',
    }

    story.append(Paragraph("SADA IPS", title_style))
    story.append(Paragraph("Plataforma de Analítica Clínica", subtitle_style))
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph(f"Reporte Clínico de Pacientes — Generado el {fecha_actual}", meta_style))
    story.append(Spacer(1, 3))

    pacientes_qs = _filtrar_pacientes(request)
    total_pacientes = pacientes_qs.count()
    pacientes_criticos = pacientes_qs.filter(es_critico=True).count()
    pacientes_hipertensos = pacientes_qs.filter(presion_sistolica__gt=140).count()
    pacientes_diabeticos = pacientes_qs.filter(glucosa__gt=126).count()

    resumen_data = [
        [
            Paragraph("<b>Total Pacientes</b>", table_cell_style),
            Paragraph(f"<font color='#7c3aed'><b>{total_pacientes}</b></font>", table_cell_style),
            Paragraph("<b>Pacientes Críticos</b>", table_cell_style),
            Paragraph(f"<font color='#be123c'><b>{pacientes_criticos}</b></font>", table_cell_style),
        ],
        [
            Paragraph("<b>Hipertensos (Sist. > 140)</b>", table_cell_style),
            Paragraph(str(pacientes_hipertensos), table_cell_style),
            Paragraph("<b>Diabéticos (Glucosa > 126)</b>", table_cell_style),
            Paragraph(str(pacientes_diabeticos), table_cell_style),
        ]
    ]

    resumen_table = Table(resumen_data, colWidths=[155, 95, 155, 95])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), gray_bg),
        ('BOX', (0,0), (-1,-1), 1, border_color),
        ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
        ('PADDING', (0,0), (-1,-1), 6),
        ('ALIGN', (1,0), (1,-1), 'CENTER'),
        ('ALIGN', (3,0), (3,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))

    story.append(Paragraph("Resumen Clínico", section_title_style))
    story.append(resumen_table)
    story.append(Spacer(1, 15))

    story.append(Paragraph(f"Listado de Pacientes ({total_pacientes} registros)", section_title_style))

    if total_pacientes == 0:
        story.append(Paragraph("No se encontraron pacientes con los filtros aplicados.", table_cell_style))
    else:
        table_data = [[
            Paragraph("ID", table_cell_header_style),
            Paragraph("Nombre Completo", table_cell_header_style),
            Paragraph("Edad", table_cell_header_style),
            Paragraph("Sexo", table_cell_header_style),
            Paragraph("IMC", table_cell_header_style),
            Paragraph("Glucosa", table_cell_header_style),
            Paragraph("P. Sistólica", table_cell_header_style),
            Paragraph("Riesgo", table_cell_header_style),
        ]]

        for p in pacientes_qs:
            nombre = f"{p.nombres} {p.apellidos}"
            imc_str = f"{p.imc:.1f}" if p.imc else "—"
            rc = RISK_COLORS.get(p.riesgo_enfermedad, '#6b6489')
            riesgo_p = Paragraph(
                f"<font color='{rc}'><b>{p.riesgo_enfermedad.upper()}</b></font>",
                table_cell_style
            )
            table_data.append([
                Paragraph(str(p.id_paciente), table_cell_style),
                Paragraph(nombre, table_cell_style),
                Paragraph(str(p.edad or '—'), table_cell_style),
                Paragraph(str(p.sexo or '—'), table_cell_style),
                Paragraph(imc_str, table_cell_style),
                Paragraph(str(p.glucosa or '—'), table_cell_style),
                Paragraph(str(p.presion_sistolica or '—'), table_cell_style),
                riesgo_p,
            ])

        pacientes_table = Table(table_data, colWidths=[35, 135, 30, 28, 65, 50, 62, 65])
        pacientes_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), purple),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, border_color),
            ('PADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, gray_bg]),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('ALIGN', (2,0), (6,-1), 'CENTER'),
        ]))
        story.append(pacientes_table)

    doc.build(story)
    buffer.seek(0)
    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_pacientes.pdf"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_csv(request):
    pacientes_qs = _filtrar_pacientes(request)
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="pacientes.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(['ID', 'Nombres', 'Apellidos', 'Edad', 'Sexo', 'Peso', 'Altura', 'IMC',
                     'Clasificacion IMC', 'Presion Sistolica', 'Presion Diastolica',
                     'Frecuencia Cardiaca', 'Glucosa', 'Colesterol', 'Saturacion O2',
                     'Temperatura', 'Antecedentes Familiares', 'Fumador', 'Consumo Alcohol',
                     'Actividad Fisica', 'Diagnostico', 'Riesgo', 'Critico', 'Fecha Consulta'])
    for p in pacientes_qs:
        writer.writerow([
            p.id_paciente, p.nombres, p.apellidos, p.edad, p.sexo, p.peso, p.altura, p.imc,
            p.clasificacion_imc, p.presion_sistolica, p.presion_diastolica,
            p.frecuencia_cardiaca, p.glucosa, p.colesterol, p.saturacion_oxigeno,
            p.temperatura, p.antecedentes_familiares, p.fumador, p.consumo_alcohol,
            p.actividad_fisica, p.diagnostico_preliminar, p.riesgo_enfermedad,
            'Si' if p.es_critico else 'No', p.fecha_consulta
        ])
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_excel(request):
    pacientes_qs = _filtrar_pacientes(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pacientes'

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4A1080', end_color='4A1080', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D2D7E5'),
        right=Side(style='thin', color='D2D7E5'),
        top=Side(style='thin', color='D2D7E5'),
        bottom=Side(style='thin', color='D2D7E5')
    )

    headers = ['ID', 'Nombres', 'Apellidos', 'Edad', 'Sexo', 'Peso', 'Altura', 'IMC',
               'Clasificacion IMC', 'Presion Sistolica', 'Presion Diastolica',
               'Frecuencia Cardiaca', 'Glucosa', 'Colesterol', 'Saturacion O2',
               'Temperatura', 'Antecedentes Familiares', 'Fumador', 'Consumo Alcohol',
               'Actividad Fisica', 'Diagnostico', 'Riesgo', 'Critico', 'Fecha Consulta']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for i, p in enumerate(pacientes_qs, 2):
        values = [
            p.id_paciente, p.nombres, p.apellidos, p.edad, p.sexo, p.peso, p.altura, p.imc,
            p.clasificacion_imc, p.presion_sistolica, p.presion_diastolica,
            p.frecuencia_cardiaca, p.glucosa, p.colesterol, p.saturacion_oxigeno,
            p.temperatura, 'Si' if p.antecedentes_familiares else 'No',
            'Si' if p.fumador else 'No', 'Si' if p.consumo_alcohol else 'No',
            p.actividad_fisica, p.diagnostico_preliminar, p.riesgo_enfermedad,
            'Si' if p.es_critico else 'No', p.fecha_consulta
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['U'].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pacientes.xlsx"'
    return response
